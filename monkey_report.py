# version: 20221108
import time
import os
import datetime
import openpyxl as xl
import sys
import shutil

# 这个会把由bluetooth导致的native_crash的文件夹全部删掉，主要是因为我们的有些板子有bluetooth的硬件问题老是报nativecrash
def delete_bluetooth():
    # return True
    return False
native_crash_bluetooth = {}

# 本地相关路径，目录
# last_log的上一级文件夹称为base_dir，传入的参数是last_log的路径
# --base_dir
# -----last_log
# -----logcat
# -----com COM
# -----monkey.log
# -----测试报告.xlsx
args = sys.argv[1:]
if len(args) != 1:
    print("usage: python %s {dir_name}" % sys.argv[0])
    sys.exit(-1)
dir_name = args[0]
if not os.path.isdir(dir_name):
    print("%s does not exist or is not a directory!" % dir_name)
    sys.exit(-1)
base_dir_path = os.path.dirname(dir_name)
xlsx_path = os.path.join(base_dir_path, "测试报告.xlsx") # 修改报告生成路径
log_share_path = os.path.join(r"\\192.168.0.195\RND Share\SW\Android\05_Log\monkey_test_Jiahao") # 修改共享信息路径
last_log_path = os.path.join(base_dir_path, "last_log") # 获取last_log文件夹路径
logcat_path = os.path.join(base_dir_path, "monkey_logcat.log") # 获取logcat路径
monkey_path = os.path.join(base_dir_path, "monkey.txt")
monkey_info_path = os.path.join(base_dir_path, "monkey_info_log.log")
monkey_error_path = os.path.join(base_dir_path, "monkey_error_log.log")
test_info_path = os.path.join(base_dir_path, "test_info.txt")
tmp_dir_path = os.path.join(base_dir_path, "tmp_log_report") # tmp提取到哪个文件夹
base_dir_name = os.path.basename(base_dir_path)
native_crash_list_path = os.path.join(tmp_dir_path, "native_crash_list.txt") # 这三个文件存放sde文件夹下的所有异常名称
crash_list_path = os.path.join(tmp_dir_path, "crash_list.txt") # 这三个文件存放sde文件夹下的所有异常名称
anr_list_path = os.path.join(tmp_dir_path, "anr_list.txt") # 这三个文件存放sde文件夹下的所有异常名称
sde_path = os.path.join(last_log_path, "sde") # 获取sde文件夹路径
sdrv_logs_path = os.path.join(last_log_path, "sdrv_logs") # sdrv_logs路径

# 主函数
def main():
    init_xlsx() # 创建xlsx报告
    write_xlsx_head() # 写报告表头
    create_tmp_dir() # 创建临时文件夹
    list_error() # 把sde下所有的错误写入到临时文件夹
    print("正在提取测试报告......")
    get_native_crash_key_info() # 读取并往报告里写入native crash的信息
    get_crash_key_info() # 读取并往报告里写入crash的信息
    get_anr_key_info() # 读取并往报告里写入anr的信息
    delete_tmp_dir() # 删除临时文件夹
    print("测试报告提取完成。\n")

# 其它通用函数
def get_date():
    today = time.strftime('%Y-%m-%d',time.localtime(time.time()))
    # return "2000-01-01"
    return today

def getYesterday(): 
    today=datetime.date.today() 
    oneday=datetime.timedelta(days=1) 
    yesterday=today-oneday  
    return str(yesterday)

# 输入error文件夹名提取其中包名，暂时没用上
def get_app_name(file_name):
    if 'native_crash' in file_name:
        app_name_rear = file_name.index("202")
        app_name = file_name[13:app_name_rear-1]
    elif 'crash' in file_name:
        app_name_rear = file_name.index("202")
        app_name = file_name[6:app_name_rear-1]
    elif 'anr' in file_name:
        app_name_rear = file_name.index("202")
        app_name = file_name[4:app_name_rear-1]
    else:
        app_name = "this is not an app"
    return app_name

# 获取sde文件夹下的文件夹名称并分类
def list_error():
    if not os.path.exists(sde_path):
        return
    # os.listdir()方法获取文件夹名字，返回数组
    file_name_list = os.listdir(sde_path)
    if not len(file_name_list) == 0:
        for file_name in file_name_list:
            a = file_name.strip()
            #app_name = get_app_name(a)
            if 'native_crash_' in file_name:
                with open(native_crash_list_path, "a", encoding="UTF_8") as f:
                    f.write(file_name+"\n")
            elif 'crash_' in file_name:
                with open(crash_list_path, "a", encoding="UTF_8") as f:
                    f.write(file_name+"\n")
            elif 'anr_' in file_name:
                with open(anr_list_path, "a", encoding="UTF_8") as f:
                    f.write(file_name+"\n")
            elif 'exceptionRecord.txt' in file_name:
                pass
            else:
                with open(tmp_dir_path+"\\others.txt", "a") as f:
                    f.write(file_name+"\n")
            #一个app只选一个文件
            # app_name = get_app_name(file_name)
            # if not app_name in app_names and not app_name == "this is not an app":
            #     app_names.append(app_name)
            #     f.write(file_name+"\n")
                #print(app_name)

def create_tmp_dir():
    if os.path.exists(tmp_dir_path):
        # 删除tmp文件夹
        try:
            shutil.rmtree(tmp_dir_path)
        except OSError as e:
            print("Error: %s - %s." % (e.filename, e.strerror))
        # print("请删除tmp文件夹后继续程序")
        # os.system("pause")
    os.makedirs(tmp_dir_path)

def delete_tmp_dir():
    if os.path.exists(tmp_dir_path):
        # 删除tmp文件夹
        try:
            shutil.rmtree(tmp_dir_path)
        except OSError as e:
            print("Error: %s - %s." % (e.filename, e.strerror))

def init_xlsx():
    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)
        print("正在删除文件夹下原有测试报告......")
    workbook = xl.Workbook()
    workbook.save(xlsx_path)

def write_xlsx_head():
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    # 设置列宽行高
    sheet.column_dimensions['A'].width=18
    sheet.column_dimensions['B'].width=200
    for i in range(1, 201):
        sheet.row_dimensions[i].height = 13.5
    # 写内容
    sheet.append(["测试日期", getYesterday()])
    sheet.append(["测试版本", "Android12Userdebug DB"])
    sheet.append(["Patch", "无"])
    sheet.append(["测试内容", "Monkey"])
    sheet.append(["测试时间", "12小时"])
    sheet.append(["log路径", log_share_path])
    # cell = sheet['B6']
    # make_hyperlink(cell, get_serial_hyperlink())
    #cell.value = '=HYPERLINK("{}", "{}")'.format("\\\\192.168.0.195\\RND Share\\SW\\Android\\05_Log\\monkey_test_Jiahao\\2022-07-19\\monkey_logcat.log", "\\\\192.168.0.195\\RND Share\\SW\\Android\\05_Log\\monkey_test_Jiahao\\2022-07-19\\monkey_logcat.log")
    sheet.append([])
    sheet.append(["测试概要"])
    workbook.save(xlsx_path)

def make_hyperlink(cell, link):
    return
    cell.hyperlink = link
    cell.value = link
    cell.style = "Hyperlink"









def get_native_crash_key_info():
    if not os.path.exists(native_crash_list_path):
        write_native_none()
        return
    native_crash_times = {}
    native_crash_content = {}
    native_crash_key = ""
    with open(native_crash_list_path, "r", encoding="UTF_8") as rf:
        for line in rf:
            native_crash_tombstone_path = os.path.join(sde_path, line.strip(), "tombstone")
            native_crash_base_path = os.path.dirname(native_crash_tombstone_path)
            native_crash_bluetooth[native_crash_base_path] = 0
            if not os.path.exists(native_crash_tombstone_path):
                print("WARNING: line %d, %s 文件不存在" % (sys._getframe().f_lineno, native_crash_tombstone_path))
                continue
            native_crash_key = get_native_crash_key(native_crash_tombstone_path)
            if native_crash_key == 0:
                continue
            if native_crash_key in native_crash_times:
                native_crash_times[native_crash_key] += 1
                continue
            else:
                native_crash_times[native_crash_key] = 1
                file = get_native_crash_content(native_crash_tombstone_path)
                native_crash_content[native_crash_key] = file.copy()
                file.clear()
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    for key in native_crash_times:
        sheet.append(["native_crash(发生%d次)" % native_crash_times[key]])
        #print(crash_content[key])
        for line in native_crash_content[key]:
            sheet.append(["", line.strip()])
    workbook.save(xlsx_path)
    if delete_bluetooth():
        for isbluetooth in native_crash_bluetooth.keys():
            # print(isbluetooth)
            # print(native_crash_bluetooth[isbluetooth])
            if native_crash_bluetooth[isbluetooth] == 1:
                try:
                    shutil.rmtree(isbluetooth)
                    print("删除native_crash_bluetooth %s" % isbluetooth)
                except OSError as e:
                    print("Error: %s - %s." % (e.filename, e.strerror))

def write_native_none():
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    sheet.append(["native_crash", "未出现"])
    workbook.save(xlsx_path)

def get_native_crash_key(native_crash_tombstone_path):
    native_crash_base_path = os.path.dirname(native_crash_tombstone_path)
    with open(native_crash_tombstone_path, "r", encoding='UTF-8') as rf:
        for line in rf:
            if 'Cmdline:' in line:
                if delete_bluetooth() and 'bluetooth' in line.lower():
                        # print(native_crash_base_path)
                        native_crash_bluetooth[native_crash_base_path] = 1
                        return 0
                return line.strip()
    return 0

def get_native_crash_content(native_crash_tombstone_path):
    write = 0
    backtrace = 0
    file = []
    with open(native_crash_tombstone_path, "r", encoding='UTF-8') as rf:
        for line in rf:
            if 'Cmdline:' in line:
                write = 1
            if 'backtrace' in line:
                backtrace = 1
            if backtrace == 1 and ( '--- ---' in line or line.strip() == ""):
                write = 0
                break
            if write == 1:
                file.append(line.strip())
    return file









def get_crash_key_info():
    if not os.path.exists(crash_list_path):
        write_crash_none()
        return
    crash_log_path = find_last_crash_log()
    with open(crash_log_path, "r", encoding='UTF-8') as rf:
        file = []
        for line in rf:
            file.append(line.strip())
        write_crash_key_info(file)

def write_crash_none():
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    sheet.append(["java_crash", "未出现"])
    workbook.save(xlsx_path)

def write_crash_key_info(file):
    # 写log
    #with open(os.path.join(tmp_dir_path, "all_crash.log"), "a", encoding="UTF_8") as wf:
    #    wf.writelines(file)
    # 写xlsx
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    write = 0
    crash_times = {}
    crash_content = {}
    last_crash = []
    crash_key = ""
    flag = ""
    for line in file:
        if write == 1:
            if flag != line[0:30]:
                write = 0
                if crash_key in crash_times:
                    crash_times[crash_key] += 1
                else:
                    crash_times[crash_key] = 1
                    crash_content[crash_key] = last_crash.copy()
                # print(crash_content[crash_key])
                last_crash.clear()
            else:
                if 'Process:' in line:
                    crash_key += line[line.find("Process:"):line.find(", PID")]
                if 'java.' in line:
                    crash_key += line[line.find("java."):len(line)]
                last_crash.append(line.strip())
        if write == 0 and 'FATAL EXCEPTION:' in line:
            crash_key = line[line.find("FATAL EXCEPTION:"):len(line)]+" "
            flag = line[0:30]
            write = 1
    if write == 1:
        write = 0
        if crash_key in crash_times:
            crash_times[crash_key] += 1
        else:
            crash_times[crash_key] = 1
            crash_content[crash_key] = last_crash.copy()
        # print(crash_content[crash_key])
        last_crash.clear()
    # location = "B" + (str)(count-2)
    # sheet[location].value = line[58:90]
    # sheet[location].font = Font(bold = True)
    # print(crash_content)
    for key in crash_times:
        sheet.append(["java_crash(发生%d次)" % crash_times[key]])
        #print(crash_content[key])
        for line in crash_content[key]:
            sheet.append(["", line.strip()])
    workbook.save(xlsx_path)

def find_last_crash_log():
    crash_log_dict = {}
    with open(crash_list_path, "r", encoding="UTF_8") as rf:
        for line in rf:
            crash_time = line.split("_")[-2] + line.split("_")[-1]
            crash_log_path = os.path.join(sde_path, line.strip(), "crash.log")
            if not os.path.exists(crash_log_path):
                print("WARNING: line %d, %s 文件不存在" % (sys._getframe().f_lineno, crash_log_path))
                continue
            crash_log_dict[crash_log_path] = crash_time
    #         crash_logs[crash_log_path] = os.path.getsize(crash_log_path)
    #         #print(os.path.getsize(crash_log_path))
    last_crash_log = max(crash_log_dict, key=crash_log_dict.get)
    print("INFO: line %d, 最新crash.log文件目录：%s" % (sys._getframe().f_lineno, last_crash_log))
    return last_crash_log

def get_anr_key_info():
    if not os.path.exists(anr_list_path):
        write_anr_none()
        return
    anr_times = {}
    anr_content = {}
    anr_key = ""
    with open(anr_list_path, "r", encoding="UTF_8") as rf:
        for line in rf:
            anr_event_log_path = os.path.join(sde_path, line.strip(), "event.log")
            if not os.path.exists(anr_event_log_path):
                print("WARNING: line %d, %s 文件不存在" % (sys._getframe().f_lineno, anr_event_log_path))
                continue

            with open(anr_event_log_path, "r", encoding='UTF-8') as evtf:
                for row in evtf:
                    if 'I am_anr' in row:
                        anr_key = row[40:65]
                        if anr_key in anr_times:
                            if row.strip() == anr_content[anr_key]:
                                continue
                            else:
                                anr_times[anr_key] += 1
                        else:
                            anr_times[anr_key] = 1
                            anr_content[anr_key] = row.strip()
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    for key in anr_times:
        sheet.append(["anr(发生%d次)" % anr_times[key], anr_content[key]])
        # print(anr_content[key])
    workbook.save(xlsx_path)

def write_anr_none():
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    sheet.append(["anr", "未出现"])
    workbook.save(xlsx_path)

if __name__ == '__main__':
    main()