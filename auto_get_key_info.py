import time
import os
import datetime
import openpyxl as xl

# 以下包含有可能要修改的路径信息
# 和测试过程中常改的信息

def is_release():
    #return True
    return False

# 本地相关路径

def get_date():
    today = time.strftime('%Y-%m-%d',time.localtime(time.time()))
    # return "2000-01-01"
    return getYesterday()
    return today

# 这是本次last_log所在的文件夹，不同版本需要根据情况修改date_dir_path
def get_date_dir_path():
    if is_release():
        return os.path.join(os.getcwd())
    else:
        return os.path.join(os.getcwd(), "logs", get_date())

# # date_dir_path
# def get_date_dir_path():
#     return os.path.join(os.getcwd())


# 获取last_log文件夹路径
def get_last_log_path():
    return os.path.join(get_date_dir_path(), "last_log")
# 获取logcat路径
def get_logcat_path():
    return os.path.join(get_date_dir_path(), "logcat.log")
def get_monkey_path():
    return os.path.join(get_date_dir_path(), "monkey.txt")
def get_monkey_info_path():
    return os.path.join(get_date_dir_path(), "monkey_info_log.log")
def get_monkey_error_path():
    return os.path.join(get_date_dir_path(), "monkey_error_log.log")
def get_test_info_path():
    return os.path.join(get_date_dir_path(), "测试信息.txt")
# tmp提取到哪个文件夹
def get_tmp_dir_path():
    return os.path.join(get_date_dir_path(), "tmp")
# 根据情况修改报告生成路径
# 压测报告存放的地址
def get_xlsx_path():
    return os.path.join(get_date_dir_path(), "测试报告.xlsx")


# 这三个文件存放sde文件夹下的所有异常名称
def get_native_crash_list_path():
    return os.path.join(get_tmp_dir_path(), "native_crash_list.txt")
def get_crash_list_path():
    return os.path.join(get_tmp_dir_path(), "crash_list.txt")
def get_anr_list_path():
    return os.path.join(get_tmp_dir_path(), "anr_list.txt")


# 获取sde文件夹路径
def get_sde_path():
    return os.path.join(get_last_log_path(), "sde")
# sdrv_logs_path
def get_sdrv_logs_path():
    return os.path.join(get_last_log_path(), "sdrv_logs")


# 更具情况更改共享信息路径
# 共享域相关路径定义
def today_log_share_path():
    if is_release():
        return ""
    else:
        return os.path.join(r"\\192.168.0.195\RND Share\SW\Android\05_Log\monkey_test_Jiahao", get_date())
# def today_log_share_path():
#     return ""
# logcat路径
def get_logcat_hyperlink():
    return os.path.join(today_log_share_path(), "logcat.log")
# monkey.log路径
def get_monkey_hyperlink():
    return os.path.join(today_log_share_path(), "monkey.txt")
# monkey_info路径
def get_monkey_info_hyperlink():
    return os.path.join(today_log_share_path(), "monkey_info_log.log")
# monkey_error路径
def get_monkey_error_hyperlink():
    return os.path.join(today_log_share_path(), "monkey_error_log.log")
# 串口信息路径
def get_serial_hyperlink():
    return os.path.join(today_log_share_path(), "[com COMxx]")
# last_log文件夹路径
def get_last_log_hyperlink():
    return os.path.join(today_log_share_path(), "last_log")
# sdrv_logs路径
def get_sdrv_logs_hyperlink():
    return os.path.join(get_last_log_hyperlink(), "sdrv_logs")
# bugreport路径
def get_bugreport_hyperlink():
    return os.path.join(today_log_share_path(), "bugreport.zip")



# 其它通用函数
def getYesterday(): 
    today=datetime.date.today() 
    oneday=datetime.timedelta(days=1) 
    yesterday=today-oneday  
    return str(yesterday)

# 输入error文件夹名提取其中包名
# 暂时没用上
def get_app_name(file_name):
    if 'native_crash' in file_name:
        app_name_rear = file_name.index("202")
        app_name = file_name[13:app_name_rear-1]
    elif 'crash' in file_name:
        app_name_rear = file_name.index("202")
        app_name = file_name[6:app_name_rear-1]
    elif 'anr' in file_name:
        app_name_rear = file_name.index("202")
        app_name = file_name[4:app_name_rear-1]
    else:
        app_name = "this is not an app"
    return app_name








# 获取sde文件夹下的文件夹名称并分类
def list_error():
    sde_path = get_sde_path()
    if not os.path.exists(sde_path):
        return
    # os.listdir()方法获取文件夹名字，返回数组
    file_name_list = os.listdir(sde_path)
    if not len(file_name_list) == 0:
        for file_name in file_name_list:
            a = file_name.strip()
            #app_name = get_app_name(a)
            if 'native_crash_' in file_name:
                with open(get_native_crash_list_path(), "a", encoding="UTF_8") as f:
                    f.write(file_name+"\n")
            elif 'crash_' in file_name:
                with open(get_crash_list_path(), "a", encoding="UTF_8") as f:
                    f.write(file_name+"\n")
            elif 'anr_' in file_name:
                with open(get_anr_list_path(), "a", encoding="UTF_8") as f:
                    f.write(file_name+"\n")
            elif 'exceptionRecord.txt' in file_name:
                pass
            else:
                with open(get_tmp_dir_path()+"\\others.txt", "a") as f:
                    f.write(file_name+"\n")
            
            #一个app只选一个文件
            # app_name = get_app_name(file_name)
            # if not app_name in app_names and not app_name == "this is not an app":
            #     app_names.append(app_name)
            #     f.write(file_name+"\n")
                #print(app_name)


def create_tmp_dir():
    tmp_dir_path = get_tmp_dir_path()
    if os.path.exists(tmp_dir_path):
        # 删除tmp文件夹
        import shutil
        try:
            shutil.rmtree(tmp_dir_path)
        except OSError as e:
            print("Error: %s - %s." % (e.filename, e.strerror))
        # print("请删除当前logs\日期文件夹下的tmp文件夹后继续程序")
        # os.system("pause")
    os.makedirs(tmp_dir_path)

def delete_tmp_dir():
    tmp_dir_path = get_tmp_dir_path()
    if os.path.exists(tmp_dir_path):
        # 删除tmp文件夹
        import shutil
        try:
            shutil.rmtree(tmp_dir_path)
        except OSError as e:
            print("Error: %s - %s." % (e.filename, e.strerror))


def auto_get_key_log():
    if not (os.path.exists(get_last_log_path()) or os.path.exists(get_logcat_path()) or os.path.exists(get_monkey_error_path()) or os.path.exists(get_monkey_path())):
        print("找不到报告，请检查脚本执行位置！")
    else:
        create_tmp_dir()
        list_error()
        create_xlsx()
        print("正在提取测试报告......")
        get_native_crash_key_info()
        get_crash_key_info()
        get_anr_key_info()
        delete_tmp_dir()
        print("测试报告提取完成。\n")
        # get_crash_key_info()
        # get_native_crash_key_info()












def get_test_info():
    test_info = []
    if os.path.exists(get_test_info_path()):
        with open(get_test_info_path(), "r", encoding="UTF_8") as f:
            for line in f:
                if line.strip() == '':
                    continue
                else:
                    test_info.append(line.strip().split("："))
    return test_info


def create_xlsx():
    xlsx_path = get_xlsx_path()
    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)
        print("正在删除文件夹下原有测试报告......")
    #print(xlsx_path)
    #print('***** 开始写入excel文件 ' + xlsx_path + ' ***** \n')
    # if os.path.exists(xlsx_path):
    #     print('***** excel已存在，在表后添加数据 ' + xlsx_path + ' ***** \n')
    #     workbook = xl.load_workbook(xlsx_path)
    # else:
    #print('创建excel ' + xlsx_path + ' ***** \n')
    workbook = xl.Workbook()
    sheet = workbook.active
    
    # 设置列宽行高
    sheet.column_dimensions['A'].width=18
    sheet.column_dimensions['B'].width=200
    for i in range(1, 201):
        sheet.row_dimensions[i].height = 13.5
    
    # 写内容
    sheet.append(["日期", get_date()])
    test_info = get_test_info()
    if not len(test_info) == 0:
        sheet.append(["测试信息"])
        for list in test_info:
            sheet.append(list)
    else:
        sheet.append(["测试信息", "无"])
    
    
    sheet.append([])
    
    if os.path.exists(get_last_log_path()):
        sheet.append(["last_log", get_last_log_hyperlink()])
        cell = sheet['B4']
        make_hyperlink(cell, get_last_log_hyperlink())
    else:
        sheet.append(["last_log", "无"])
    
    if os.path.exists(get_logcat_path()):
        sheet.append(["logcat", get_logcat_hyperlink()])
        cell = sheet['B5']
        make_hyperlink(cell, get_logcat_hyperlink())
    else:
        sheet.append(["logcat", "无"])
    
    serial_hyperlink =  "无"
    for filename in os.listdir(get_date_dir_path()):
        if "com com" in filename.lower() or "com_com" in filename.lower():
            serial_hyperlink = get_serial_hyperlink()
            break
    sheet.append(["串口信息", serial_hyperlink])
    cell = sheet['B6']
    make_hyperlink(cell, get_serial_hyperlink())


    if os.path.exists(get_monkey_path()):
        sheet.append(["monkey.txt", get_monkey_hyperlink()])
        cell = sheet['B7']
        make_hyperlink(cell, get_monkey_hyperlink())

    if os.path.exists(get_monkey_info_path()):
        sheet.append(["monkey_info", get_monkey_info_hyperlink()])
        cell = sheet['B7']
        make_hyperlink(cell, get_monkey_info_hyperlink())
    
    if os.path.exists(get_monkey_error_path()):
        sheet.append(["monkey_error", get_monkey_error_hyperlink()])
        cell = sheet['B8']
        make_hyperlink(cell, get_monkey_error_hyperlink())
    #cell.value = '=HYPERLINK("{}", "{}")'.format("\\\\192.168.0.195\\RND Share\\SW\\Android\\05_Log\\monkey_test_Jiahao\\2022-07-19\\logcat.log", "\\\\192.168.0.195\\RND Share\\SW\\Android\\05_Log\\monkey_test_Jiahao\\2022-07-19\\logcat.log")
    
    if os.path.exists(get_sdrv_logs_path()):
        sheet.append(["sdrv_logs", get_sdrv_logs_hyperlink()])
        cell = sheet['B9']
        make_hyperlink(cell, get_sdrv_logs_hyperlink())
    else:
        sheet.append(["sdrv_logs", "无"])
    
    bugreport_hyperlink =  "无"
    for filename in os.listdir(get_date_dir_path()):
        if "bugreport" in filename.lower():
            bugreport_hyperlink = get_bugreport_hyperlink()
            break
    sheet.append(["bugreport", bugreport_hyperlink])
    cell = sheet['B10']
    make_hyperlink(cell, get_bugreport_hyperlink())
    
    sheet.append([])
    sheet.append(["测试概要"])
    
    workbook.save(xlsx_path)
    #print('***** 生成Excel文件 ' + xlsx_path + ' ***** \n')


def make_hyperlink(cell, link):
    return
    cell.hyperlink = link
    cell.value = link
    cell.style = "Hyperlink"
    









def abc():
    pass
def get_native_crash_key_info():
    if not os.path.exists(get_native_crash_list_path()):
        write_native_none()
        return
    
    native_crash_times = {}
    native_crash_content = {}
    native_crash_key = ""
    
    with open(get_native_crash_list_path(), "r", encoding="UTF_8") as rf:
        for line in rf:
            native_crash_tombstone_path = os.path.join(get_sde_path(), line.strip(), "tombstone")
            if not os.path.exists(native_crash_tombstone_path):
                print("WARNING: %s 文件不存在" % native_crash_tombstone_path)
                continue
            
            # # xlsx标题
            # xlsx_path = get_xlsx_path()
            # workbook = xl.load_workbook(xlsx_path)
            # sheet = workbook.active
            # sheet.append(["native_crash"])
            # workbook.save(xlsx_path)
            
            native_crash_key = get_native_crash_key(native_crash_tombstone_path)
            if native_crash_key in native_crash_times:
                native_crash_times[native_crash_key] += 1
                continue
            else:
                native_crash_times[native_crash_key] = 1
                file = get_native_crash_content(native_crash_tombstone_path)
                native_crash_content[native_crash_key] = file.copy()
                file.clear()
    
    xlsx_path = get_xlsx_path()
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    for key in native_crash_times:
        sheet.append(["native_crash(发生%d次)" % native_crash_times[key]])
        #print(crash_content[key])
        for line in native_crash_content[key]:
            sheet.append(["", line.strip()])

    workbook.save(xlsx_path)


def write_native_none():
    xlsx_path = get_xlsx_path()
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    sheet.append(["native_crash", "未出现"])
    workbook.save(xlsx_path)
    

def get_native_crash_key(native_crash_tombstone_path):
    with open(native_crash_tombstone_path, "r", encoding='UTF-8') as rf:
        for line in rf:
            if 'Cmdline:' in line:
                return line.strip()

def get_native_crash_content(native_crash_tombstone_path):
    write = 0
    backtrace = 0
    file = []
    with open(native_crash_tombstone_path, "r", encoding='UTF-8') as rf:
        for line in rf:
            if 'Cmdline:' in line:
                write = 1
            if 'backtrace' in line:
                backtrace = 1
            if backtrace == 1 and ( '--- ---' in line or line.strip() == ""):
                write = 0
                break
            if write == 1:
                file.append(line.strip())
    return file









def get_crash_key_info():
    if not os.path.exists(get_crash_list_path()):
        write_crash_none()
        return
    crash_log_path = find_last_crash_log()
    with open(crash_log_path, "r", encoding='UTF-8') as rf:
        file = []
        for line in rf:
            file.append(line.strip())
        write_crash_key_info(file)


def write_crash_none():
    xlsx_path = get_xlsx_path()
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    sheet.append(["java_crash", "未出现"])
    workbook.save(xlsx_path)


def write_crash_key_info(file):
    # 写log
    #with open(os.path.join(get_tmp_dir_path(), "all_crash.log"), "a", encoding="UTF_8") as wf:
    #    wf.writelines(file)
    
    # 写xlsx
    del file[0]
    xlsx_path = get_xlsx_path()
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    write = 1
    crash_times = {}
    crash_content = {}
    last_crash = []
    crash_key = ""
    for line in file:
        if 'FATAL EXCEPTION' in line:
            if write == 1:
                crash_content[crash_key] = last_crash.copy()
            # print(crash_content[crash_key])
            last_crash.clear()
            continue
        elif 'Process:' in line:
            crash_key = line[49:80]
            if crash_key in crash_times:
                crash_times[crash_key] += 1
                write = 0
            else:
                crash_times[crash_key] = 1
                last_crash.append(line.strip())
                write = 1
            continue
        if write == 1:
            last_crash.append(line.strip())
            # location = "B" + (str)(count-2)
            # sheet[location].value = line[58:90]
            # sheet[location].font = Font(bold = True)
    # print(crash_content)
    if write == 1:
        crash_content[crash_key] = last_crash.copy()
    last_crash.clear()
    for key in crash_times:
        sheet.append(["java_crash(发生%d次)" % crash_times[key]])
        #print(crash_content[key])
        for line in crash_content[key]:
            sheet.append(["", line.strip()])

    workbook.save(xlsx_path)


def find_last_crash_log():
    crash_logs = {}
    with open(get_crash_list_path(), "r", encoding="UTF_8") as rf:
        for line in rf:
            crash_log_path = os.path.join(get_sde_path(), line.strip(), "crash.log")
            if not os.path.exists(crash_log_path):
                print("WARNING: %s 文件不存在" % crash_log_path)
                continue
            crash_logs[crash_log_path] = os.path.getsize(crash_log_path)
            #print(os.path.getsize(crash_log_path))
    last_crash_log = max(crash_logs, key=crash_logs.get)
    # print("INFO: 最新crash.log文件目录：%s" % last_crash_log)
    return last_crash_log












def get_anr_key_info():
    if not os.path.exists(get_anr_list_path()):
        write_anr_none()
        return
    with open(get_anr_list_path(), "r", encoding="UTF_8") as rf:
        for line in rf:
            anr_event_log_path = os.path.join(get_sde_path(), line.strip(), "event.log")
            if not os.path.exists(anr_event_log_path):
                print("WARNING: %s 文件不存在" % anr_event_log_path)
                continue

            with open(anr_event_log_path, "r", encoding='UTF-8') as rf:
                for line in rf:
                    if 'I am_anr' in line:
                        write_anr_key_info(line)


def write_anr_none():
    xlsx_path = get_xlsx_path()
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    sheet.append(["anr", "未出现"])
    workbook.save(xlsx_path)
    
    
def write_anr_key_info(line):
    # 写log
    #with open(os.path.join(get_key_info_dir_path(), "all_anr.log"), "a", encoding="UTF_8") as wf:
    #    wf.write(line)
    
    # 写xlsx
    xlsx_path = get_xlsx_path()
    workbook = xl.load_workbook(xlsx_path)
    sheet = workbook.active
    sheet.append(["anr", line.strip()])
    workbook.save(xlsx_path)
    



    

if __name__ == '__main__':
    auto_get_key_log()